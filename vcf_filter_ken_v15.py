import pandas
import re
import os
import pickle
import csv
import argparse
import gzip
import openpyxl
import math
from openpyxl.utils import get_column_letter
import statistics
import numpy as np
import concurrent.futures
from collections import defaultdict


	# Line  23165: ENSG00000164362,60251,ENST00000484238,1283010,1287086,4076
	# Line  23166: ENSG00000164362,60504,ENST00000310581,1295068,1318453,23385
	# Line  23167: ENSG00000164362,60512,ENST00000656021,1295047,1318453,23406
	# Line  23168: ENSG00000164362,60552,ENST00000460137,1294989,1318453,23464
	# Line  23169: ENSG00000164362,60564,ENST00000334602,1294989,1318453,23464
	# Line  23170: ENSG00000164362,60572,ENST00000667927,1264534,1264592,58
	# Line  23171: ENSG00000164362,60578,ENST00000503656,1266524,1266535,11
exclude_list=["chr5_1253832","chr5_1254369","chr5_1254506","chr5_1255288","chr5_1255411","chr5_1258598","chr5_1258658","chr5_1260473","chr5_1280339","chr5_1282430"]


parser = argparse.ArgumentParser()
parser.add_argument('--input_st', help='input filename')
parser.add_argument('--input_nd', help='input filename')
parser.add_argument('--label_st', help='input filename')
parser.add_argument('--label_nd', help='input filename')
parser.add_argument('--chr_pos', help='input filename')
parser.add_argument('--check_DB', help='input filename')
parser.add_argument('--two_pass', help='input filename')
parser.add_argument('--merge_DB', help='input filename')
parser.add_argument('--mark_CFS', help='mark CFS based on MiDASeq hg38')
parser.add_argument('--mark_RT', help='mark RT based on RT_A2780')
parser.add_argument('--comment', help='comment line start with')
parser.add_argument('--filter', help='remove PON and polyG or A')
parser.add_argument('--header', help='header line start with')
parser.add_argument('--include_string', help='include string')
parser.add_argument('--drop', help='drop header') #output file without header
parser.add_argument('--output', help='output filename') #output file without header
parser.add_argument('--delimiter', help='delimiter (default=\\t)') #output file without header
parser.add_argument('--exclude', help='output the file with string excluded') #output file without header
parser.add_argument('--second_extract', help='extract other line based on output column (based on header)') #output file without header
args = parser.parse_args()




CFS_list_P1 = {}
with open("Supplementary_Table_1.csv", mode="rt", encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        chrom = row['Chr']
        if chrom not in CFS_list_P1:
            CFS_list_P1[chrom] = []
        CFS_list_P1[chrom].append({
            'start': int(row['Start']), 
            'end': int(row['End']), 
            'id': row['cCFS ID']
        })

# Load Priority 2: FRA_Summary_Table.csv 
CFS_list_P2 = {}
with open("Supplementary_Table_2.csv", mode="rt", encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        chrom = row['Chr']
        if chrom not in CFS_list_P2:
            CFS_list_P2[chrom] = []
        CFS_list_P2[chrom].append({
            'start': int(row['Start']), 
            'end': int(row['End']), 
            'id': row['mCFS ID']
        })

def vcf_to_pandas(header="#", delimiter="\t", input_arr=""):
    chr_prog = re.compile('chr[0-9XYxy]+$')
    output_arr=[]
    chr_position_list={}
    header_list=[]
    for i,l in enumerate(input_arr):
        if l.startswith(header): #CHROM	POS	ID	REF	ALT	QUAL	FILTER	INFO	FORMAT	RPE1_CON.sv_prep.sorted.bam
            l=l.replace(header,"")
            header_list=l.split(delimiter)
            #print(header_list)
        else:
            output_dict={}
            if len(header_list):
                l_list=l.split(delimiter)
                for j, k in enumerate(l_list):
                    output_dict[header_list[j]]=k
                if re.match(chr_prog, output_dict["CHROM"]):
                    if output_dict["CHROM"] not in chr_position_list:
                        chr_position_list[output_dict["CHROM"]]=[]
                    if output_dict["POS"] not in chr_position_list[output_dict["CHROM"]]:
                        chr_position_list[output_dict["CHROM"]].append(output_dict["POS"])                    
                    output_arr.append(output_dict)
    output_arr_df = pandas.DataFrame.from_records(output_arr)
    return [output_arr,output_arr_df, chr_position_list]

def list_to_chr_position_ET(input_arr):
    output_arr=[]
    chr_position_list={}
    position_index=0
    position_index_list=[]
    for l in input_arr:
        if l["CHROM"] not in chr_position_list:
            chr_position_list[l["CHROM"]]={}
            chr_position_list[l["CHROM"]]["POS"]=[]
            chr_position_list[l["CHROM"]]["POS_idx"]=[]
        if l["POS"] not in chr_position_list[l["CHROM"]]["POS"]:
            chr_position_list[l["CHROM"]]["POS"].append(l["POS"])
            chr_position_list[l["CHROM"]]["POS_idx"].append(position_index)
            position_index=position_index+1
    return chr_position_list
                
def string_filter(skp_header="",delimiter="",input_arr="",include_string="",search_index=""):
    output_arr=[]
    sk_output_arr=[]
    total_line=0
    line_keep=0
    line_skip=0
    for i,l in enumerate(input_arr):
        total_line=total_line+1
        if i==0:
            if skp_header:
                pass
            else:
                output_arr.append(l)
                sk_output_arr.append(l)
        else:
            if search_index:
                if include_string==l.split(delimiter)[search_index]:
                    output_arr.append(l)
            else:
                if include_string in l:
                    line_keep=line_keep+1
                    output_arr.append(l)
                else:
                    line_skip=line_skip+1
                    sk_output_arr.append(l)
    return {'output_arr':output_arr,'sk_output_arr':sk_output_arr}
    
def vcf_input(input_fn="", output="",drop="",second_extract="", delimiter="", comment="",write_output=""):
    if input_fn:# and input_nd:# and include_string:
        input_arr=[]
        if drop:
            skp_header=1
        else:
            skp_header=0
        if output:
            output_fn=output
        else:
            fn=input_fn.split(".")[0]
            fn_ex=input_fn.split(".")[-1]
            output_fn=fn+"_output."+fn_ex
        print("processing "+input_fn+" to "+output_fn)
        if delimiter:
            delimiter=delimiter
        else:   delimiter="\t"
        cwd = os.getcwd()
        if fn_ex=="gz":
            with gzip.open(input_fn, mode="rt") as f:
                file_content = f.read().splitlines()
                if comment: pass
                else: comment="##"
                for l in file_content:
                    if comment:
                        if l.startswith(comment):
                            pass
                        else:
                            input_arr.append(l)
                    else:
                        input_arr.append(l)
        else:
            with open(input_fn, 'r') as Tfile:
                lines = Tfile.read().splitlines()
                for l in lines:
                    input_arr.append(l)
        
        output=vcf_to_pandas(input_arr=input_arr)
        return output

        #string_filter_output=string_filter(skp_header=skp_header,include_string=include_string,delimiter=delimiter,input_arr=input_arr)

        #Stripped.Sequence
        if second_extract:
            second_extract_output=[]
            second_extract=0
            #get header
            header_tmp=input_arr[0]
            header=header_tmp.split(delimiter)
            search_index=""
            for i, h in enumerate(header):
                if h==second_extract:
                    search_index=i
            if search_index:
                for s in string_filter_output['output_arr'][(skp_header*-1)+1:]:
                    s_search_string=s.split(delimiter)[search_index]
                    second_extract_output_tmp=string_filter(skp_header=1,include_string=s_search_string,delimiter=delimiter,input_arr=string_filter_output['sk_output_arr'],search_index=search_index)
                    second_extract_output=second_extract_output+second_extract_output_tmp['output_arr']
            else:
                print("cannot find "+second_extract+" from the header")
        if write_output:#write output
            if exclude: output_write=string_filter_output['sk_output_arr']#write exclude
            else:   output_write=string_filter_output['output_arr']
            with open(cwd+'/'+output_fn, 'w') as Wfile:
                for l in output_write:
                    Wfile.write(str(l)+"\n")
                Wfile.close()
            if second_extract:
                with open(cwd+'/'+output_fn+"second_extract", 'w') as Wfile:
                    for l in second_extract_output:
                        Wfile.write(str(l)+"\n")
                    Wfile.close()
        
    else:
        parser.print_help()

def df_vcf(pandas_df, fn): #first_output_df.to_csv("first_stpass_output_temp.txt", sep='\t', encoding='utf-8', index=False, header=True)
    pandas_df.to_csv(fn, sep='\t', encoding='utf-8', index=False, header=True)
    with open(fn,'r') as f:
        fcounter=0
        for line in f:
            if fcounter:
                newf=newf+line.strip()+"\n"
            else:   newf="#"+line.strip()+"\n"
            fcounter=fcounter+1
        f.close()
    with open(fn,'w') as f:
        f.write(newf)
        f.close()

def process_rt_chunk(chunk_list, mark_RT_flag, rt_files, rt_a_dict):
    """Worker function to process a chunk of variants for Replication Timing."""
    chunk_output = []
    RT_cutoff = 200001    
    map_RT = {}
    
    for i, j in enumerate(chunk_list):
        if mark_RT_flag:
            for RT_file in rt_files:
                j[RT_file] = ""
                
        chromosome = j['CHROM']
        chromosome_pos = int(j['POS'])
        
        for RT_file in rt_a_dict:
            RT_A_list = rt_a_dict[RT_file]
            if j['CHROM'] in RT_A_list:
                RTA_Chr = RT_A_list[j['CHROM']]
                RTA_Chr_check = 1
                RTA = ""
                
                for iii, RTS in enumerate(RTA_Chr):
                    if RTS['Coordinate'] >= chromosome_pos and RTA_Chr_check: 
                        #compare the distance of iii-1 vs iii
                        if iii > 0:
                            if (RTS['Coordinate']-chromosome_pos) <= (chromosome_pos-RTA_Chr[iii-1]['Coordinate']) and (RTS['Coordinate']-chromosome_pos)<=RT_cutoff:   
                                RTA=RTS['RT']
                            if (RTS['Coordinate']-chromosome_pos) > (chromosome_pos-RTA_Chr[iii-1]['Coordinate']) and (chromosome_pos-RTA_Chr[iii-1]['Coordinate'])<=RT_cutoff:   
                                RTA=RTA_Chr[iii-1]['RT']                                    
                        else:
                            #this is first element
                            if (RTS['Coordinate']-chromosome_pos)<=RT_cutoff:   
                                RTA=RTS['RT']
                                
                        RTA_Chr_check = 0                                    
                        if len(RTA):
                            j[RT_file] = float(RTA)
                            if RT_file not in map_RT:
                                map_RT[RT_file] = 1
                            else: 
                                map_RT[RT_file] = map_RT[RT_file] + 1
                        else:
                            j[RT_file] = RTA
                            
        chunk_output.append(j)
        
    return chunk_output

def add_rt(first_input_RT):
    firstpass_output_tmp = []
    
    # 1. Determine the chunk size (Let's split it into 8 chunks for processing)
    # Ensure chunk_size is at least 1 to prevent division by zero
    num_chunks = 8 
    chunk_size = math.ceil(len(first_input_RT) / num_chunks)
    if chunk_size == 0: 
        chunk_size = 1
        
    # 2. Slice the main list into a list of smaller chunks
    chunks = [first_input_RT[i:i + chunk_size] for i in range(0, len(first_input_RT), chunk_size)]
    
    # 3. Process the chunks concurrently
    with concurrent.futures.ProcessPoolExecutor() as executor:
        futures = []
        
        # Submit each chunk along with the necessary variables to prevent Windows pickling errors
        for chunk in chunks:
            future = executor.submit(process_rt_chunk, chunk, args.mark_RT, RT_filelist, RT_A_list_dict)
            futures.append(future)
            
        # 4. Gather the results sequentially to maintain the original list order
        for future in futures:
            firstpass_output_tmp.extend(future.result())
            
    return firstpass_output_tmp


def zero_pass(first_input,remove_mate=1, filter="PON;maxPolyGLength;maxPolyAHomLength;smallDelInsertionArtifact", chrom="1", QUAL=200, position=0):
    chrlen={'chr1':248956422,'chr2':242193529,'chr3':198295559,'chr4':190214555,'chr5':181538259,'chr6':170805979,'chr7':159345973,'chr8':145138636,'chr9':138394717,'chr10':133797422,'chr11':135086622,'chr12':133275309,'chr13':114364328,'chr14':107043718,'chr15':101991189,'chr16':90338345,'chr17':83257441,'chr18':80373285,'chr19':58617616,'chr20':64444167,'chr21':46709983,'chr22':50818468,'chrX':156040895,'chrY':57227415,'chr1_cen_start':122503247,'chr2_cen_start':92188145,'chr3_cen_start':91553419,'chr4_cen_start':49712061,'chr5_cen_start':47153439,'chr6_cen_start':58553888,'chr7_cen_start':58169653,'chr8_cen_start':44033744,'chr9_cen_start':43389635,'chr10_cen_start':39686682,'chr11_cen_start':51078348,'chr12_cen_start':34769407,'chr13_cen_start':16000000,'chr14_cen_start':16000000,'chr15_cen_start':17083673,'chr16_cen_start':36337666,'chr17_cen_start':22813679,'chr18_cen_start':15460899,'chr19_cen_start':24498980,'chr20_cen_start':26436232,'chr21_cen_start':10864560,'chr22_cen_start':12954788,'chrX_cen_start':58605579,'chrY_cen_start':10316944,'chr1_cen_end':124849129,'chr2_cen_end':94090557,'chr3_cen_end':93655574,'chr4_cen_end':51743951,'chr5_cen_end':50059807,'chr6_cen_end':59829934,'chr7_cen_end':61528020,'chr8_cen_end':45877265,'chr9_cen_end':45518558,'chr10_cen_end':41497440,'chr11_cen_end':54425074,'chr12_cen_end':37185252,'chr13_cen_end':18051248,'chr14_cen_end':18173523,'chr15_cen_end':19725254,'chr16_cen_end':38265669,'chr17_cen_end':26616164,'chr18_cen_end':20861206,'chr19_cen_end':27190874,'chr20_cen_end':30038348,'chr21_cen_end':12915808,'chr22_cen_end':15054318,'chrX_cen_end':62412542,'chrY_cen_end':10544039}    
    first_input_list=first_input
    if position==0:
        position_abs=1500000
    else:
        position_abs=position
        
    firstpass_output_tmp=[]
    firstpass_output=[]
    prog = re.compile('.*?MATEID=([a-zA-Z_0-9]+).*')
    prog_et= re.compile('.*?EVENTTYPE=([a-zA-Z]+).*')
    chr_prog = re.compile('chr[0-9XYxy]+$')
    filter_arr = set(filter.split(";"))
    mateid_to_remove=[]
    filter_flag=0
    if len(filter_arr): filter_flag=1
    for i,j in enumerate(first_input_list):
        #check second input
        j['EVENTTYPE']=""
        chromosome=j['CHROM']
        if re.match(chr_prog, j["CHROM"]):
            chrom_flag=1
        else: chrom_flag=0
        if chrom_flag:
            chromosome_pos=int(j['POS'])
            if chromosome_pos<=chrlen[chromosome+"_cen_start"]: #
                j['POS_Chr']=round(chromosome_pos/(chrlen[chromosome]-chrlen[chromosome+"_cen_start"])*100,4) ##
                j['POS_Chr_abs']=round(chromosome_pos/position_abs,2) ##
            elif chromosome_pos>=chrlen[chromosome+"_cen_end"]: #:
                j['POS_Chr']=round(((chrlen[chromosome]-chrlen[chromosome+"_cen_end"])-(chromosome_pos-chrlen[chromosome+"_cen_end"]))/(chrlen[chromosome]-chrlen[chromosome+"_cen_end"])*100,4)
                j['POS_Chr_abs']=round(((chrlen[chromosome]-chrlen[chromosome+"_cen_end"])-(chromosome_pos-chrlen[chromosome+"_cen_end"]))/position_abs,2)
            else:
                j['POS_Chr']=100
                j['POS_Chr_abs']=round(chromosome_pos/position_abs,2) ##
            if re.match(prog_et, j["INFO"]):
                EVENTTYPE=re.search(prog_et, j["INFO"])                
                j['EVENTTYPE']=EVENTTYPE.group(1)
            if filter_flag: #skip if filter match
                filter_check=1
                j_filter_arr=j["FILTER"].split(";")
                for jj in j_filter_arr:
                    if jj in filter_arr:    filter_check=0
                if float(j["QUAL"])<QUAL: filter_check=0
                if filter_check and chrom_flag:    
                    j["QUAL"]=round(float(j["QUAL"]),2)
                    if re.match(prog, j["INFO"]):
                        mateid=re.search(prog, j["INFO"])
                        j["mateid"]=mateid.group(1)
                    else:   j["mateid"]=""
                    firstpass_output_tmp.append(j)
                else:
                    mateid=""
                    if re.match(prog, j["INFO"]):
                        mateid=re.search(prog, j["INFO"])
                        j["mateid"]=mateid.group(1)
                        mateid_to_remove.append(mateid.group(1))                
            else:
                firstpass_output_tmp.append(j)
    if remove_mate:
        for i,j in enumerate(firstpass_output_tmp):
            if j['ID'] in mateid_to_remove:
                pass
                #firstpass_output.append(j)
            else:
                firstpass_output.append(j)
        return firstpass_output
    else:
        return firstpass_output_tmp
        
def process_chr_position_lst(chr_position_lst):
    """Helper function to run the POS extraction loop concurrently."""
    db_tmp = []
    db_tmp_index = []
    
    for i in chr_position_lst:
        chr_t = i
        for ijk, j in enumerate(chr_position_lst[i]["POS"]):
            db_tmp.append(chr_t + "_" + j)
            db_tmp_index.append(chr_position_lst[i]["POS_idx"][ijk])
            
    return db_tmp, db_tmp_index

def first_pass(first_input_A, remove_mate=1, mark_chr=1, mark_CFS=1, position=0, QUAL=200, ignore_ET=1):
    chrlen={'chr1':248956422,'chr2':242193529,'chr3':198295559,'chr4':190214555,'chr5':181538259,'chr6':170805979,'chr7':159345973,'chr8':145138636,'chr9':138394717,'chr10':133797422,'chr11':135086622,'chr12':133275309,'chr13':114364328,'chr14':107043718,'chr15':101991189,'chr16':90338345,'chr17':83257441,'chr18':80373285,'chr19':58617616,'chr20':64444167,'chr21':46709983,'chr22':50818468,'chrX':156040895,'chrY':57227415,'chr1_cen_start':122503247,'chr2_cen_start':92188145,'chr3_cen_start':91553419,'chr4_cen_start':49712061,'chr5_cen_start':47153439,'chr6_cen_start':58553888,'chr7_cen_start':58169653,'chr8_cen_start':44033744,'chr9_cen_start':43389635,'chr10_cen_start':39686682,'chr11_cen_start':51078348,'chr12_cen_start':34769407,'chr13_cen_start':16000000,'chr14_cen_start':16000000,'chr15_cen_start':17083673,'chr16_cen_start':36337666,'chr17_cen_start':22813679,'chr18_cen_start':15460899,'chr19_cen_start':24498980,'chr20_cen_start':26436232,'chr21_cen_start':10864560,'chr22_cen_start':12954788,'chrX_cen_start':58605579,'chrY_cen_start':10316944,'chr1_cen_end':124849129,'chr2_cen_end':94090557,'chr3_cen_end':93655574,'chr4_cen_end':51743951,'chr5_cen_end':50059807,'chr6_cen_end':59829934,'chr7_cen_end':61528020,'chr8_cen_end':45877265,'chr9_cen_end':45518558,'chr10_cen_end':41497440,'chr11_cen_end':54425074,'chr12_cen_end':37185252,'chr13_cen_end':18051248,'chr14_cen_end':18173523,'chr15_cen_end':19725254,'chr16_cen_end':38265669,'chr17_cen_end':26616164,'chr18_cen_end':20861206,'chr19_cen_end':27190874,'chr20_cen_end':30038348,'chr21_cen_end':12915808,'chr22_cen_end':15054318,'chrX_cen_end':62412542,'chrY_cen_end':10544039}    
    prog_DEL= re.compile(r'.*?(chr[a-zA-Z_0-9]+):(\d+).*')
    prog_BND_target= re.compile('.*?(chr[a-zA-Z_0-9]+).*')
    if position==0:
        position_abs=1500000
    else:
        position_abs=position
    first_input_chr_position_lst=list_to_chr_position_ET(first_input_A[0])    
    second_input_chr_position_lst=list_to_chr_position_ET(first_input_A[1])
    
# Use concurrent futures to process A and B at the exact same time
    with concurrent.futures.ProcessPoolExecutor(max_workers=2) as executor:
        # Submit both tasks to the CPU cores
        future_A = executor.submit(process_chr_position_lst, first_input_chr_position_lst)
        future_B = executor.submit(process_chr_position_lst, second_input_chr_position_lst)
        
        # Wait for them to finish and unpack the results
        db_A_tmp, db_A_tmp_index = future_A.result()
        db_B_tmp, db_B_tmp_index = future_B.result()
    
    db_B_dict = {}
    for xyz, y in enumerate(db_B_tmp):
        db_B_dict[y] = first_input_A[1][db_B_tmp_index[xyz]]["EVENTTYPE"]

    # OPTIMIZATION: Use a Set to store common entries (prevents duplicates automatically and makes lookups instant)
    db_common_set = set()
    
    for ijk, x in enumerate(db_A_tmp):
        if x in db_B_dict: # O(1) instant lookup
            if first_input_A[0][db_A_tmp_index[ijk]]["EVENTTYPE"] == db_B_dict[x] or ignore_ET:
                db_common_set.add(x)
                
    for kkk in exclude_list:
        db_common_set.add(kkk)
        
    mateid_to_remove_lst={}
    output_tmp_lst={}
    firstpass_output={}
    for i in [0,1]:
        mateid_to_remove_lst[i]=[]
        output_tmp_lst[i]=[]
        input_list=first_input_A[i]
        for k, j in enumerate(input_list):
            if j['CHROM']+"_"+j['POS'] in db_common_set:
                    mateid_to_remove_lst[i].append(j["mateid"])
            else:
                if mark_CFS:
                    j['CFS'] = 0
                    j['mCFS'] = 0
                    j['cCFS'] = 0
                    j['cCFS ID'] = ""
                    j['mCFS ID'] = ""                    
                    j['BNDtarget'] = ""
                    j['DEL_size'] = ""
                    j['CFS_DEL_size'] = ""
                    j['mCFS_DEL_size'] = ""
                    j['cCFS_DEL_size'] = ""
                    
                    
                    var_pos = int(j['POS'])
                    found_fra_id = None

                    # Tier 1: Check Common_Fragile_Sites (Priority)
                    if j['CHROM'] in CFS_list_P1:
                        for ref in CFS_list_P1[j['CHROM']]:
                            if ref['start'] <= var_pos <= ref['end']:
                                j['cCFS ID'] = ref['id']
                                j['CFS'] = 1
                                j['cCFS'] = 1
                                break # Stop searching Priority 1 if found

                    # Tier 2: Check FRA_Summary_Table (Fallback)
                    if not found_fra_id and j['CHROM'] in CFS_list_P2:
                        for ref in CFS_list_P2[j['CHROM']]:
                            if ref['start'] <= var_pos <= ref['end']:
                                j['mCFS ID'] = ref['id']
                                j['CFS'] = 1
                                j['mCFS'] = 1
                                break # Stop searching Fallback if found

                if j['EVENTTYPE']=="BND":
                    if re.match(prog_BND_target, j["ALT"]):
                        BND_target=re.search(prog_BND_target, j["ALT"])                
                        j['BNDtarget']=BND_target.group(1)                    
                if j['EVENTTYPE']=="DEL":
                    DEL_size=0
                    #calculate DEL size
                    if re.match(prog_DEL, j["ALT"]): #chr1:6385836
                        DEL_target=re.search(prog_DEL, j["ALT"])                
                        if DEL_target.group(2)>=j['POS']:
                            DEL_size=int(DEL_target.group(2))-int(j['POS'])
                        else:
                            DEL_size=int(j['POS'])-int(DEL_target.group(2))
                        j['DEL_size']=DEL_size
                        if j['cCFS'] and DEL_size:    j['cCFS_DEL_size']=DEL_size
                        if j['mCFS'] and DEL_size:    j['mCFS_DEL_size']=DEL_size
                        if j['CFS'] and DEL_size:    j['CFS_DEL_size']=DEL_size
                output_tmp_lst[i].append(j)

    for i in [0,1]:
        firstpass_output[i]=[]
        if remove_mate:
            for k,j in enumerate(output_tmp_lst[i]):
                if j['ID'] in mateid_to_remove_lst[i]:
                    pass
                else:
                    firstpass_output[i].append(j)
            firstpass_output[i]=add_rt(firstpass_output[i])
        else:
            firstpass_output[i]=add_rt(output_tmp_lst[i])
    
    return firstpass_output
    
def export_to_excel(input_obj, input_dict, cfs="CFS"): #{"QUAL_list":QUAL_list, "chrlen":chrlen, "EVENTTYPE_list":EVENTTYPE_list,"input_type":"zeropass"}
    QUAL_list=input_dict["QUAL_list"]
    chrlen=input_dict["chrlen"]
    EVENTTYPE_list=input_dict["EVENTTYPE_list"]
    input_type=input_dict["input_type"]
    
    if input_type=="zeropass": #input_obj is a list input list
        A_zeropass_df = pandas.DataFrame.from_records(input_obj[0])
        B_zeropass_df = pandas.DataFrame.from_records(input_obj[1])
        output_list=[]
        output_list.append(",,,SGL,DEL,BND,DUP,INV,INS,,"+cfs+",SGL,DEL,BND,DUP,INV,INS,,DEL(log10),<=1,<=2,<=3,<=4,<=5,<=6,>6")
        for q in QUAL_list:
            first_Qa=A_zeropass_df[A_zeropass_df['QUAL']>=q]
            second_Qa=B_zeropass_df[B_zeropass_df['QUAL']>=q]
            #first
            output_row="Qual:,"+str(q)+","+firstlabel+","
            for et in EVENTTYPE_list:
                output_row=output_row+str(len(first_Qa[first_Qa['EVENTTYPE']==et]))+","
            output_list.append(output_row[:-1])
            #second
            output_row=","+str(q)+","+secondlabel+","
            for et in EVENTTYPE_list:
                output_row=output_row+str(len(second_Qa[second_Qa['EVENTTYPE']==et]))+","
            output_list.append(output_row[:-1])
        output_list.append("")
        output_list.append("")
        output_list.append("")

    if input_type=="firstpass": #input_obj is a list input list
        A_zeropass_df = pandas.DataFrame.from_records(input_obj[0])
        B_zeropass_df = pandas.DataFrame.from_records(input_obj[1])

        output_list=[]
        header=",,,SGL,DEL,BND,DUP,INV,INS,,"+cfs+",SGL,DEL,BND,DUP,INV,INS,,DEL(log10),<=1,<=2,<=3,<=4,<=5,<=6,>6"
        for RT_file in RT_AFbin_dict:
            header=header+",,"+RT_file+",1 late,2,3,4,5,6 early"
        header=header+",,CFS DEL(log10),<=1,<=2,<=3,<=4,<=5,<=6,>6,,mCFS DEL(log10),<=1,<=2,<=3,<=4,<=5,<=6,>6,,cCFS DEL(log10),<=1,<=2,<=3,<=4,<=5,<=6,>6"
        output_list.append(header)
        for q in QUAL_list:
            Q_Arr=[]
            f_Qa=A_zeropass_df[A_zeropass_df['QUAL']>=q]
            s_Qa=B_zeropass_df[B_zeropass_df['QUAL']>=q]
            Q_Arr.append([f_Qa,firstlabel])
            Q_Arr.append([s_Qa,secondlabel])

            for first_Qa in Q_Arr:
                #first
                output_row="Qual:,"+str(q)+","+first_Qa[1]+","
                first_Qa=first_Qa[0]
                for et in EVENTTYPE_list:
                    first_Qa_tmp=""
                    first_Qa_tmp=first_Qa[first_Qa['EVENTTYPE']==et]
                    #output_row=output_row+str(len(first_Qa[first_Qa['EVENTTYPE']==et]))+","
                    output_row=output_row+str(len(first_Qa_tmp))+","
                output_row=output_row+","
                output_row=output_row+","
                for et in EVENTTYPE_list:
                    first_Qa_tmp=""
                    first_Qa_tmp=first_Qa[first_Qa['EVENTTYPE']==et]
                    output_row=output_row+str(len(first_Qa_tmp[first_Qa_tmp[cfs]==1]))+","
                output_row=output_row+","
                output_row=output_row+","
        ###export del size
                del_arr=[]
                first_Qa_tmp=""
                first_Qa_tmp=first_Qa[first_Qa['EVENTTYPE']=="DEL"]
                del_tmp=0
                for del_s in DELsize_list:
                    del_arr.append(len(first_Qa_tmp[first_Qa_tmp['DEL_size']<=10**del_s])-del_tmp)
                    del_tmp=len(first_Qa_tmp[first_Qa_tmp['DEL_size']<=10**del_s])
                del_arr.append(len(first_Qa_tmp[first_Qa_tmp['DEL_size']>10**DELsize_list[-1]]))
                for del_s in del_arr:
                    output_row=output_row+str(del_s)+","
        ###export del size end

                output_row=output_row+","
                output_row=output_row+","

                for RT_file in RT_AFbin_dict:
            ###export RTA2780        
                    del_arr=[]
                    first_Qa_tmp=""
                    first_Qa_tmp=first_Qa[first_Qa[RT_file].notnull()]
                    first_Qa_tmp[RT_file] = pandas.to_numeric(first_Qa_tmp[RT_file], errors='coerce')

                    #print(first_Qa_tmp)
                    del_tmp=0
                    for RT_s in RT_AFbin_dict[RT_file]:
                        del_arr.append(len(first_Qa_tmp[first_Qa_tmp[RT_file]<=RT_s])-del_tmp)
                        del_tmp=len(first_Qa_tmp[first_Qa_tmp[RT_file]<=RT_s])
                    del_arr.append(len(first_Qa_tmp[first_Qa_tmp[RT_file]>RT_AFbin_dict[RT_file][-1]]))
                    for RT_s in del_arr:
                        output_row=output_row+str(RT_s)+","
                    output_row=output_row+","
                    output_row=output_row+","
            ###export RTA2780 end
            
###export CFS del size 
                del_arr=[]
                first_Qa_tmp=""
                # ADD .copy() AND to_numeric() HERE
                first_Qa_tmp = first_Qa[first_Qa['EVENTTYPE']=="DEL"].copy()
                first_Qa_tmp['CFS_DEL_size'] = pandas.to_numeric(first_Qa_tmp['CFS_DEL_size'], errors='coerce')
                
                del_tmp=0
                for del_s in DELsize_list:
                    del_arr.append(len(first_Qa_tmp[first_Qa_tmp['CFS_DEL_size']<=10**del_s])-del_tmp)
                    del_tmp=len(first_Qa_tmp[first_Qa_tmp['CFS_DEL_size']<=10**del_s])
                del_arr.append(len(first_Qa_tmp[first_Qa_tmp['CFS_DEL_size']>10**DELsize_list[-1]]))
                for del_s in del_arr:
                    output_row=output_row+str(del_s)+","
        ###export del size end

                output_row=output_row+","
                output_row=output_row+","

###export mCFS del size 
                del_arr=[]
                first_Qa_tmp=""
                # ADD .copy() AND to_numeric() HERE
                first_Qa_tmp = first_Qa[first_Qa['EVENTTYPE']=="DEL"].copy()
                first_Qa_tmp['mCFS_DEL_size'] = pandas.to_numeric(first_Qa_tmp['mCFS_DEL_size'], errors='coerce')
                
                del_tmp=0
                for del_s in DELsize_list:
                    del_arr.append(len(first_Qa_tmp[first_Qa_tmp['mCFS_DEL_size']<=10**del_s])-del_tmp)
                    del_tmp=len(first_Qa_tmp[first_Qa_tmp['mCFS_DEL_size']<=10**del_s])
                del_arr.append(len(first_Qa_tmp[first_Qa_tmp['mCFS_DEL_size']>10**DELsize_list[-1]]))
                for del_s in del_arr:
                    output_row=output_row+str(del_s)+","
        ###export del size end

                output_row=output_row+","
                output_row=output_row+","

###export cCFS del size 
                del_arr=[]
                first_Qa_tmp=""
                # ADD .copy() AND to_numeric() HERE
                first_Qa_tmp = first_Qa[first_Qa['EVENTTYPE']=="DEL"].copy()
                first_Qa_tmp['cCFS_DEL_size'] = pandas.to_numeric(first_Qa_tmp['cCFS_DEL_size'], errors='coerce')
                
                del_tmp=0
                for del_s in DELsize_list:
                    del_arr.append(len(first_Qa_tmp[first_Qa_tmp['cCFS_DEL_size']<=10**del_s])-del_tmp)
                    del_tmp=len(first_Qa_tmp[first_Qa_tmp['cCFS_DEL_size']<=10**del_s])
                del_arr.append(len(first_Qa_tmp[first_Qa_tmp['cCFS_DEL_size']>10**DELsize_list[-1]]))
                for del_s in del_arr:
                    output_row=output_row+str(del_s)+","
        ###export del size end

                output_row=output_row+","
                output_row=output_row+","
            
                output_list.append(output_row[:-1])
            
        output_list.append("")
        output_list.append("")
        output_list.append("")

#######################################
    if input_type=="firstpass": EVENTTYPE_list=input_dict["EVENTTYPE_list"]

    for n in range(1,23):
        output_list.append("Chr"+str(n)+",,,SGL,DEL,BND,DUP,INV,INS")
        first_Ca=A_zeropass_df[A_zeropass_df['CHROM']=="chr"+str(n)]
        second_Ca=B_zeropass_df[B_zeropass_df['CHROM']=="chr"+str(n)]
        for q in QUAL_list:
            first_Qa=first_Ca[first_Ca['QUAL']>=q]
            second_Qa=second_Ca[second_Ca['QUAL']>=q]
            #first
            output_row="Qual:,"+str(q)+","+firstlabel+","
            for et in EVENTTYPE_list:
                output_row=output_row+str(len(first_Qa[first_Qa['EVENTTYPE']==et]))+","
            output_list.append(output_row[:-1])
            #second
            output_row=","+str(q)+","+secondlabel+","
            for et in EVENTTYPE_list:
                output_row=output_row+str(len(second_Qa[second_Qa['EVENTTYPE']==et]))+","
            output_list.append(output_row[:-1])
        output_list.append("")
        output_list.append("")
        output_list.append("")    
    #print(first_input[2]['chr1'])
    #print(second_input[1])
    EVENTTYPE_list_ALL=EVENTTYPE_list
    EVENTTYPE_list_ALL.insert(0,"ALL")
    for et in EVENTTYPE_list_ALL:
        output_list.append(et)
        RCP_header="Relative Chr position,,,"
        for RCP in range(0,110,10):
            RCP_header=RCP_header+str(RCP)+","
        #ALL
        output_list.append(RCP_header)
        for q in QUAL_list:
            if et=="ALL":
                firstRa=A_zeropass_df
                secondRa=B_zeropass_df
            else:
                firstRa=A_zeropass_df[A_zeropass_df['EVENTTYPE']==et]
                secondRa=B_zeropass_df[B_zeropass_df['EVENTTYPE']==et]
            first_Qa=firstRa[firstRa['QUAL']>=q]
            second_Qa=secondRa[secondRa['QUAL']>=q]
            #first
            output_row="Qual:,"+str(q)+","+firstlabel+","
            for RCP in range(0,110,10):
                output_row=output_row+str(len(first_Qa[first_Qa['POS_Chr']<=RCP+10])-len(first_Qa[first_Qa['POS_Chr']<=RCP]))+","
            output_list.append(output_row[:-1])
            #second
            output_row=","+str(q)+","+secondlabel+","
            for RCP in range(0,110,10):
                output_row=output_row+str(len(second_Qa[second_Qa['POS_Chr']<=RCP+10])-len(second_Qa[second_Qa['POS_Chr']<=RCP]))+","
            output_list.append(output_row[:-1])
        output_list.append("")
        output_list.append("")
        output_list.append("")

    for et in EVENTTYPE_list_ALL:
        output_list.append(et)
        RCP_header="Abs Chr position("+str(abs_pos)+"),,,"
        for RCP in range(0,105,5):
            RCP_header=RCP_header+str(RCP)+","
        #ALL
        output_list.append(RCP_header)
        for q in QUAL_list:
            if et=="ALL":
                firstRa=A_zeropass_df
                secondRa=B_zeropass_df
            else:
                firstRa=A_zeropass_df[A_zeropass_df['EVENTTYPE']==et]
                secondRa=B_zeropass_df[B_zeropass_df['EVENTTYPE']==et]
            first_Qa=firstRa[firstRa['QUAL']>=q]
            second_Qa=secondRa[secondRa['QUAL']>=q]
            #first
            output_row="Qual:,"+str(q)+","+firstlabel+","
            for RCP in range(0,105,5):
                output_row=output_row+str(len(first_Qa[first_Qa['POS_Chr_abs']<=RCP+10])-len(first_Qa[first_Qa['POS_Chr_abs']<=RCP]))+","
            output_list.append(output_row[:-1])
            #second
            output_row=","+str(q)+","+secondlabel+","
            for RCP in range(0,105,5):
                output_row=output_row+str(len(second_Qa[second_Qa['POS_Chr_abs']<=RCP+10])-len(second_Qa[second_Qa['POS_Chr_abs']<=RCP]))+","
            output_list.append(output_row[:-1])
        output_list.append("")
        output_list.append("")
        output_list.append("")

        
    with open(cwd+'/output.csv', 'w') as Wfile:
        for l in output_list:
            Wfile.write(l.strip()+"\n")
        
############################
    if input_type=="firstpass":
        EVENTTYPE_list=input_dict["EVENTTYPE_list"]
        A_zeropass_df = pandas.DataFrame.from_records(input_obj[0])
        B_zeropass_df = pandas.DataFrame.from_records(input_obj[1])
        output_list=[]

        for et in input_dict["EVENTTYPE_list"]:        
            if et!="ALL":
                header=et+","
                for RT_file in RT_AFbin_dict:
                    header=header+",,"+RT_file+",1 late,2,3,4,5,6 early"
                output_list.append(header)
                for q in QUAL_list:
                    Q_Arr=[]
                    f_Qa=A_zeropass_df[A_zeropass_df['QUAL']>=q]
                    s_Qa=B_zeropass_df[B_zeropass_df['QUAL']>=q]
                    Q_Arr.append([f_Qa,firstlabel])
                    Q_Arr.append([s_Qa,secondlabel])
                    for first_Qa in Q_Arr:
                        #first
                        output_row="Qual:,"+str(q)+","+first_Qa[1]+",,"
                        first_Qa=first_Qa[0]                  
                        for RT_file in RT_AFbin_dict:
                    ###export RTA2780        
                            del_arr=[]
                            first_Qa_tmp=""
                            first_Qa_tmp=first_Qa[first_Qa['EVENTTYPE']==et]
                            first_Qa_tmp=first_Qa_tmp[first_Qa_tmp[RT_file].notnull()]
                            first_Qa_tmp[RT_file] = pandas.to_numeric(first_Qa_tmp[RT_file], errors='coerce')

                            #print(first_Qa_tmp)
                            del_tmp=0
                            for RT_s in RT_AFbin_dict[RT_file]:
                                del_arr.append(len(first_Qa_tmp[first_Qa_tmp[RT_file]<=RT_s])-del_tmp)
                                del_tmp=len(first_Qa_tmp[first_Qa_tmp[RT_file]<=RT_s])
                            del_arr.append(len(first_Qa_tmp[first_Qa_tmp[RT_file]>RT_AFbin_dict[RT_file][-1]]))
                            for RT_s in del_arr:
                                output_row=output_row+str(RT_s)+","
                            output_row=output_row+","
                            output_row=output_row+","
                        output_list.append(output_row[:-1])
                output_list.append("")
                output_list.append("")
                output_list.append("")            
            
    with open(cwd+'/output_details.csv', 'w') as Wfile:
        for l in output_list:
            Wfile.write(l.strip()+"\n")
            
    templatefile=cwd+'/output_summary_v2.xlsx'
    output_path = cwd+'/'+input_type+'-output_summary - '+firstlabel+' vs '+secondlabel+'.xlsx'
    if os.path.isfile(output_path):
        templatefile=output_path

    workbook = openpyxl.load_workbook(filename=templatefile)
    worksheet_output = workbook["output ("+cfs+")"]

    with open(cwd+'/output.csv') as f:
        reader = csv.reader(f, delimiter=',')
        for row_index, row in enumerate(reader):
            for column_index, cell in enumerate(row):
                if cell.isnumeric():
                    cell=int(cell)
                column_letter = get_column_letter((column_index + 1))
                worksheet_output['%s%s'%(column_letter, (row_index + 1))].value = cell

############################
    if input_type=="firstpass":
    ###export deletion
        A_zeropass_df = pandas.DataFrame.from_records(input_obj[0])
        B_zeropass_df = pandas.DataFrame.from_records(input_obj[1])
        for et in input_dict["EVENTTYPE_list"]:
            if et!="ALL":
                Q_Arr=[]        
                f_Qa=A_zeropass_df[A_zeropass_df['EVENTTYPE']==et]
                s_Qa=B_zeropass_df[B_zeropass_df['EVENTTYPE']==et]
                Q_Arr.append([f_Qa,firstlabel])
                Q_Arr.append([s_Qa,secondlabel])        
                for first_Qa in Q_Arr:
                    label=first_Qa[1]
                    first_Qa=first_Qa[0]
                    first_Qa.to_csv(label+"_"+et+"_out.csv", index=False)


##copy
    if input_type=="firstpass":
        worksheet_output = workbook["more ("+cfs+")"]

        with open(cwd+'/output_details.csv') as f:
            reader = csv.reader(f, delimiter=',')
            for row_index, row in enumerate(reader):
                for column_index, cell in enumerate(row):
                    if cell.isnumeric():
                        cell=int(cell)
                    column_letter = get_column_letter((column_index + 1))
                    worksheet_output['%s%s'%(column_letter, (row_index + 1))].value = cell    

    if input_type=="firstpass":
        Q_Arr=[firstlabel,secondlabel]
        for et in input_dict["EVENTTYPE_list"]:
            if et!="ALL":
                for label in Q_Arr:
                    worksheet_output = workbook.create_sheet(title=label+"_"+et)
                    worksheet_output = workbook[label+"_"+et]
                    with open(cwd+'/'+label+"_"+et+"_out.csv") as f:
                        reader = csv.reader(f, delimiter=',')
                        for row_index, row in enumerate(reader):
                            for column_index, cell in enumerate(row):
                                if cell.isnumeric():
                                    cell=int(cell)
                                column_letter = get_column_letter((column_index + 1))
                                worksheet_output['%s%s'%(column_letter, (row_index + 1))].value = cell
                                
    workbook.save(input_type+'-output_summary - '+firstlabel+' vs '+secondlabel+'.xlsx')

        


if __name__ == '__main__':

    #load RT list from file
    RT_filelist={"H3K4me3_RPE1":"H3K4me3_RPE1_hg38.txt","H3K27me3_RPE1":"H3K27me3_RPE1_hg38.txt","H3K27Ac_RPE1":"H3K27Ac_RPE1_hg38.txt", "RT_A2780":"RT_A2780_hg38_smoothed.txt","RT_GM12878":"RT_GM12878_hg38_smoothed.txt","RT_HCC1143":"RT_HCC1143_hg38_smoothed.txt","RT_HCC1954":"RT_HCC1954_hg38_smoothed.txt","RT_HEK293T":"RT_HEK293T_hg38_smoothed.txt","RT_RPE1":"RT_RPE1_hg38_smoothed.txt","RT_RPE1_2019":"RT_RPE1-2019_hg38_smoothed.txt"}
    RT_AFbin_dict={}
    RT_A_list_dict={}
    DELsize_list=[1,2,3,4,5,6]
    firstlabel=args.label_st
    secondlabel=args.label_nd
    QUAL_list=[100, 150, 200, 250, 300]
    chrlen={'chr1':248956422,'chr2':242193529,'chr3':198295559,'chr4':190214555,'chr5':181538259,'chr6':170805979,'chr7':159345973,'chr8':145138636,'chr9':138394717,'chr10':133797422,'chr11':135086622,'chr12':133275309,'chr13':114364328,'chr14':107043718,'chr15':101991189,'chr16':90338345,'chr17':83257441,'chr18':80373285,'chr19':58617616,'chr20':64444167,'chr21':46709983,'chr22':50818468,'chrX':156040895,'chrY':57227415,'chr1_cen_start':122503247,'chr2_cen_start':92188145,'chr3_cen_start':91553419,'chr4_cen_start':49712061,'chr5_cen_start':47153439,'chr6_cen_start':58553888,'chr7_cen_start':58169653,'chr8_cen_start':44033744,'chr9_cen_start':43389635,'chr10_cen_start':39686682,'chr11_cen_start':51078348,'chr12_cen_start':34769407,'chr13_cen_start':16000000,'chr14_cen_start':16000000,'chr15_cen_start':17083673,'chr16_cen_start':36337666,'chr17_cen_start':22813679,'chr18_cen_start':15460899,'chr19_cen_start':24498980,'chr20_cen_start':26436232,'chr21_cen_start':10864560,'chr22_cen_start':12954788,'chrX_cen_start':58605579,'chrY_cen_start':10316944,'chr1_cen_end':124849129,'chr2_cen_end':94090557,'chr3_cen_end':93655574,'chr4_cen_end':51743951,'chr5_cen_end':50059807,'chr6_cen_end':59829934,'chr7_cen_end':61528020,'chr8_cen_end':45877265,'chr9_cen_end':45518558,'chr10_cen_end':41497440,'chr11_cen_end':54425074,'chr12_cen_end':37185252,'chr13_cen_end':18051248,'chr14_cen_end':18173523,'chr15_cen_end':19725254,'chr16_cen_end':38265669,'chr17_cen_end':26616164,'chr18_cen_end':20861206,'chr19_cen_end':27190874,'chr20_cen_end':30038348,'chr21_cen_end':12915808,'chr22_cen_end':15054318,'chrX_cen_end':62412542,'chrY_cen_end':10544039}
    EVENTTYPE_list=["SGL","DEL","BND","DUP","INV","INS"]



    if args.chr_pos:
        chr_pos_set=int(args.chr_pos)
        abs_pos=chr_pos_set
    else:
        chr_pos_set=0
        abs_pos=1500000

    for RT_file in RT_filelist:
        RT_A_list={}
        RT_A_arr=[]
        arr1 = np.array([])
        with open(RT_filelist[RT_file], mode="rt") as f:
            file_content = f.read().splitlines()
            for l in file_content:
                if not l.startswith("#"):
                    l_list=l.split("\t")
                    l_list[0]="chr"+str(l_list[0])
                    if "." not in l_list[1]:
                        if l_list[0] in RT_A_list:
                            if len(l_list[2]) and l_list[2] != "NaN":
                                RT_A_list[l_list[0]].append({'Coordinate':int(l_list[1]),'RT':l_list[2]})
                                RT_A_arr.append(float(l_list[2]))
                        else:
                            RT_A_list[l_list[0]]=[]
                            if len(l_list[2]) and l_list[2] != "NaN":
                                RT_A_list[l_list[0]].append({'Coordinate':int(l_list[1]),'RT':l_list[2]})
                                RT_A_arr.append(float(l_list[2]))


        RT_AFbin=[]
        RT_AFbin.append((np.percentile(np.array(RT_A_arr), 16)))
        RT_AFbin.append((np.percentile(np.array(RT_A_arr), 32)))
        RT_AFbin.append((np.percentile(np.array(RT_A_arr), 48)))
        RT_AFbin.append((np.percentile(np.array(RT_A_arr), 64)))
        RT_AFbin.append((np.percentile(np.array(RT_A_arr), 80)))
        
        RT_AFbin_dict[RT_file]=RT_AFbin
        RT_A_list_dict[RT_file]=RT_A_list
        total_RT_position=0
        for chromsome in RT_A_list:
            total_RT_position=total_RT_position+len(RT_A_list[chromsome])
        
        print(RT_file+"contains "+str(total_RT_position)+" Replication timing Coordinate")
        b_tmp=""
        for b in RT_AFbin:
            b_tmp=b_tmp+str(b)+","
        print("with following bins: "+b_tmp[:-1])



    cwd = os.getcwd()
    first_input = vcf_input(input_fn=args.input_st)
    second_input = vcf_input(input_fn=args.input_nd)

    # 1. Process the Two Input Files Concurrently
    with concurrent.futures.ProcessPoolExecutor(max_workers=8) as executor:
        future_A = executor.submit(zero_pass, first_input[0], QUAL=0, position=chr_pos_set)
        future_B = executor.submit(zero_pass, second_input[0], QUAL=0, position=chr_pos_set)
        
        zero_pass_output_A = future_A.result()
        zero_pass_output_B = future_B.result()

    export_to_excel([zero_pass_output_A,zero_pass_output_B], {"QUAL_list":QUAL_list, "chrlen":chrlen, "EVENTTYPE_list":EVENTTYPE_list,"input_type":"zeropass"})

    EVENTTYPE_list=["SGL","DEL","BND","DUP","INV","INS"]
    
    first_pass_output = first_pass([zero_pass_output_A,zero_pass_output_B], QUAL=0, position=chr_pos_set)

    export_to_excel(first_pass_output, {"QUAL_list":QUAL_list, "chrlen":chrlen, "EVENTTYPE_list":EVENTTYPE_list,"input_type":"firstpass"})
    
    EVENTTYPE_list=["SGL","DEL","BND","DUP","INV","INS"]
    export_to_excel(first_pass_output, {"QUAL_list":QUAL_list, "chrlen":chrlen, "EVENTTYPE_list":EVENTTYPE_list,"input_type":"firstpass"},"mCFS")
    
    EVENTTYPE_list=["SGL","DEL","BND","DUP","INV","INS"]
    export_to_excel(first_pass_output, {"QUAL_list":QUAL_list, "chrlen":chrlen, "EVENTTYPE_list":EVENTTYPE_list,"input_type":"firstpass"},"cCFS")
    
    print("Run Completed")